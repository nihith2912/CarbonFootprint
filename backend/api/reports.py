import os
import csv
from decimal import Decimal
from datetime import datetime, timedelta
from django.conf import settings
from django.db.models import Sum
from .models import CarbonEntry, Recommendation, Report
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


def get_reporting_data(user, range_type):
    """
    Retrieves carbon footprint stats for a user over a specific timeframe.
    """
    today = datetime.now().date()
    if range_type == 'Daily':
        start_date = today
    elif range_type == 'Weekly':
        start_date = today - timedelta(days=7)
    elif range_type == 'Monthly':
        start_date = today - timedelta(days=30)
    else:  # Annual
        start_date = today - timedelta(days=365)

    entries = CarbonEntry.objects.filter(user=user, date__gte=start_date).order_by('-date')
    
    # Calculate statistics
    total_co2 = entries.aggregate(total=Sum('emissions_co2'))['total'] or Decimal('0.0')
    category_breakdown = list(
        entries.values('category')
        .annotate(total=Sum('emissions_co2'))
        .order_by('-total')
    )
    
    # Comparison against average user (standard benchmark: 400kg monthly, scaled to period)
    scale_factor = 1.0
    if range_type == 'Daily':
        scale_factor = 1.0 / 30.0
    elif range_type == 'Weekly':
        scale_factor = 7.0 / 30.0
    elif range_type == 'Annual':
        scale_factor = 12.0
    
    avg_user_co2 = Decimal('400.0') * Decimal(str(scale_factor))
    reduction_percentage = 0.0
    if avg_user_co2 > 0:
        reduction_percentage = float((avg_user_co2 - total_co2) / avg_user_co2) * 100.0

    # Trees saved: 1 tree absorbs ~22kg of CO2 per year (1.83kg per month)
    trees_saved = float(avg_user_co2 - total_co2) / 22.0 if total_co2 < avg_user_co2 else 0.0
    if trees_saved < 0:
        trees_saved = 0.0

    return {
        'entries': entries,
        'total_co2': total_co2,
        'category_breakdown': category_breakdown,
        'avg_user_co2': avg_user_co2,
        'reduction_percentage': round(reduction_percentage, 1),
        'trees_saved': round(trees_saved, 2),
        'start_date': start_date,
        'end_date': today,
    }


def generate_pdf_report(user, range_type, data, filepath):
    """
    Creates a styled PDF report using reportlab.
    """
    doc = SimpleDocTemplate(filepath, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    story = []
    
    styles = getSampleStyleSheet()
    
    # Custom Styles
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=24,
        textColor=colors.HexColor('#065f46'), # Emerald Green
        spaceAfter=15
    )
    subtitle_style = ParagraphStyle(
        'SubtitleStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=12,
        textColor=colors.HexColor('#4b5563'),
        spaceAfter=25
    )
    h2_style = ParagraphStyle(
        'H2Style',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=16,
        textColor=colors.HexColor('#0f766e'),
        spaceBefore=15,
        spaceAfter=10
    )
    body_style = ParagraphStyle(
        'BodyStyle',
        parent=styles['BodyText'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor('#1f2937'),
        leading=14
    )

    # Document Header
    story.append(Paragraph("EcoTrack AI - Sustainability Report", title_style))
    story.append(Paragraph(
        f"Generated for: <b>{user.username}</b> ({user.email}) | "
        f"Period: {range_type} ({data['start_date']} to {data['end_date']})", 
        subtitle_style
    ))
    
    # Executive Summary Cards
    story.append(Paragraph("Executive Summary", h2_style))
    summary_text = (
        f"During this period, your total carbon footprint was <b>{data['total_co2']:.2f} kg CO₂</b>. "
        f"Compared to the average user benchmark of <b>{data['avg_user_co2']:.2f} kg CO₂</b>, you achieved a "
        f"<b>{data['reduction_percentage']}%</b> {'reduction' if data['reduction_percentage'] >= 0 else 'increase'} in emissions. "
        f"This is equivalent to planting <b>{data['trees_saved']}</b> mature trees!"
    )
    story.append(Paragraph(summary_text, body_style))
    story.append(Spacer(1, 15))

    # Category Breakdown Table
    story.append(Paragraph("Emissions by Category", h2_style))
    table_data = [['Category', 'Total Emissions (kg CO₂)', 'Percentage of Total']]
    
    for cat in data['category_breakdown']:
        cat_name = cat['category'].title()
        cat_total = cat['total']
        percentage = (cat_total / data['total_co2'] * 100) if data['total_co2'] > 0 else Decimal('0.0')
        table_data.append([cat_name, f"{cat_total:.2f}", f"{percentage:.1f}%"])
        
    if len(table_data) == 1:
        table_data.append(['No entries recorded', '0.00', '0%'])

    t = Table(table_data, colWidths=[200, 150, 150])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#065f46')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f3f4f6')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
    ]))
    story.append(t)
    story.append(Spacer(1, 20))

    # Detailed Entries Log
    story.append(Paragraph("Activity Log", h2_style))
    log_data = [['Date', 'Category', 'Inputs Details', 'Emissions (kg CO₂)']]
    for entry in data['entries'][:15]: # Show latest 15 entries in PDF
        input_str = ", ".join([f"{k}: {v}" for k, v in entry.inputs.items()])
        log_data.append([
            str(entry.date),
            entry.category.title(),
            input_str[:40] + ('...' if len(input_str) > 40 else ''),
            f"{entry.emissions_co2:.2f}"
        ])
    
    if len(log_data) == 1:
        log_data.append(['-', '-', 'No data recorded', '0.00'])

    t_log = Table(log_data, colWidths=[80, 80, 240, 100])
    t_log.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f766e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
        ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
    ]))
    story.append(t_log)
    story.append(Spacer(1, 20))

    # Recommendations Footer
    story.append(Paragraph("Key Recommendations for You", h2_style))
    recommendations = Recommendation.objects.filter(category_type=data['category_breakdown'][0]['category'])[:2] if data['category_breakdown'] else Recommendation.objects.all()[:2]
    
    for rec in recommendations:
        story.append(Paragraph(f"• <b>{rec.title}</b>: {rec.description} (Potential Savings: <b>{rec.co2_savings_kg} kg CO₂</b>)", body_style))
        story.append(Spacer(1, 5))

    doc.build(story)


def generate_excel_report(user, range_type, data, filepath):
    """
    Creates a formatted Excel spreadsheet using openpyxl.
    """
    wb = Workbook()
    
    # 1. Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.views.sheetView[0].showGridLines = True

    # Palette styles
    green_fill = PatternFill(start_color="065F46", end_color="065F46", fill_type="solid")
    teal_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    light_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    
    font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_normal = Font(name="Calibri", size=11)

    # Title block
    ws_summary.merge_cells("A1:D2")
    title_cell = ws_summary["A1"]
    title_cell.value = f"EcoTrack AI Carbon Footprint Report ({range_type})"
    title_cell.font = font_title
    title_cell.fill = green_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Metrics
    ws_summary["A4"] = "User"
    ws_summary["B4"] = user.username
    ws_summary["A5"] = "Period"
    ws_summary["B5"] = f"{data['start_date']} to {data['end_date']}"
    ws_summary["A6"] = "Total Carbon (kg CO₂)"
    ws_summary["B6"] = data['total_co2']
    ws_summary["A7"] = "Reduction vs Avg User"
    ws_summary["B7"] = f"{data['reduction_percentage']}%"
    ws_summary["A8"] = "Trees Saved Equivalent"
    ws_summary["B8"] = data['trees_saved']

    for row in range(4, 9):
        ws_summary[f"A{row}"].font = font_bold
        ws_summary[f"B{row}"].font = font_normal

    # Category Breakdown Headers
    ws_summary["A11"] = "Category Breakdown"
    ws_summary["A11"].font = Font(name="Calibri", size=13, bold=True, color="065F46")

    ws_summary["A12"] = "Category"
    ws_summary["B12"] = "Total CO2 (kg)"
    ws_summary["A12"].font = font_header
    ws_summary["A12"].fill = teal_fill
    ws_summary["B12"].font = font_header
    ws_summary["B12"].fill = teal_fill

    row_idx = 13
    for cat in data['category_breakdown']:
        ws_summary.cell(row=row_idx, column=1, value=cat['category'].title()).font = font_normal
        ws_summary.cell(row=row_idx, column=2, value=cat['total']).font = font_normal
        row_idx += 1

    # 2. Entries Sheet
    ws_entries = wb.create_sheet(title="All Entries")
    ws_entries.views.sheetView[0].showGridLines = True

    headers = ["Date", "Category", "CO2 Emissions (kg)", "Input Parameters"]
    for col_num, header in enumerate(headers, 1):
        cell = ws_entries.cell(row=1, column=col_num, value=header)
        cell.font = font_header
        cell.fill = green_fill
        cell.alignment = Alignment(horizontal="center")

    for row_num, entry in enumerate(data['entries'], 2):
        ws_entries.cell(row=row_num, column=1, value=str(entry.date)).font = font_normal
        ws_entries.cell(row=row_num, column=2, value=entry.category.title()).font = font_normal
        ws_entries.cell(row=row_num, column=3, value=float(entry.emissions_co2)).font = font_normal
        ws_entries.cell(row=row_num, column=4, value=json.dumps(entry.inputs)).font = font_normal

    # Adjust columns widths
    for ws in [ws_summary, ws_entries]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(filepath)


def generate_csv_report(user, data, filepath):
    """
    Creates a simple flat CSV export of carbon entries.
    """
    with open(filepath, mode='w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['User', user.username, 'Email', user.email])
        writer.writerow(['Generated At', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        writer.writerow([])
        writer.writerow(['Entry Date', 'Category', 'Emissions CO2 (kg)', 'Input Values JSON'])
        
        for entry in data['entries']:
            writer.writerow([
                entry.date,
                entry.category.title(),
                entry.emissions_co2,
                json.dumps(entry.inputs)
            ])


def create_user_report(user, range_type, format_type):
    """
    High-level trigger that gathers data, generates requested report,
    records the file metadata in the Report model, and returns its URL.
    """
    # 1. Gather data
    data = get_reporting_data(user, range_type)
    
    # 2. Create reports folder if not exists
    reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
    os.makedirs(reports_dir, exist_ok=True)
    
    # 3. Create filename
    filename = f"EcoTrack_{range_type}_{user.username}_{datetime.now().strftime('%Y%m%d%H%M%S')}"
    ext = format_type.lower()
    if ext == 'excel':
        ext = 'xlsx'
    
    filename = f"{filename}.{ext}"
    filepath = os.path.join(reports_dir, filename)
    file_url = os.path.join(settings.MEDIA_URL, 'reports', filename).replace('\\', '/')

    # 4. Generate report based on type
    if format_type == 'PDF':
        generate_pdf_report(user, range_type, data, filepath)
    elif format_type == 'Excel':
        generate_excel_report(user, range_type, data, filepath)
    elif format_type == 'CSV':
        generate_csv_report(user, data, filepath)
        
    # 5. Save report DB entry
    report_db = Report.objects.create(
        user=user,
        title=f"{range_type} Carbon Footprint Summary ({format_type})",
        range_type=range_type,
        format_type=format_type,
        file_path=file_url
    )
    
    return report_db
